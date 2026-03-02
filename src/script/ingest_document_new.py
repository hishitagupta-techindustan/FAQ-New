


"""
Structured FAQ Ingestion Script
PDF → Chunked Text → LLM (per chunk) → Merge JSON → Vector DB
"""

import sys
import json
from pathlib import Path
from loguru import logger
import fitz
from typing import List, Optional, Dict
from pydantic import BaseModel, Field

from langchain_openai import ChatOpenAI
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_core.messages import HumanMessage, SystemMessage

# Add src to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from config import settings
from script.vectorstore import VectorStore
from concurrent.futures import ThreadPoolExecutor, as_completed


# =====================================================
# ---------------- Pydantic Schema --------------------
# =====================================================

class FAQItem(BaseModel):
    question: str = Field(..., min_length=5, max_length=200)
    question_variations: List[str] = Field(..., min_items=2, max_items=3)
    


class Topic(BaseModel):
    topic_id: str
    title: str
    faqs: List[FAQItem] = Field(..., min_items=3, max_items=3)


class StructuredFAQ(BaseModel):
    topics: List[Topic] = Field(..., min_items=3, max_items=5)


# =====================================================
# ---------------- PDF Extraction ---------------------
# =====================================================

def extract_full_text(pdf_path: Path) -> str:
    """Extract full cleaned text from PDF, preserving structure."""
    doc = fitz.open(pdf_path)
    pages_text = []

    for page_num, page in enumerate(doc):
        page_text = page.get_text()
        if page_text.strip():
            pages_text.append(f"[PAGE {page_num + 1}]\n{page_text}")

    doc.close()

    full_text = "\n\n".join(pages_text)
    return full_text


def extract_keywords_from_text(text: str) -> List[str]:
    """Extract meaningful keywords from text for later validation."""
    import re
    # Normalize whitespace but keep structure
    words = re.findall(r'\b[A-Za-z][a-zA-Z\-]{3,}\b', text)
    # Deduplicate preserving order, lowercase
    seen = set()
    keywords = []
    for w in words:
        lw = w.lower()
        if lw not in seen:
            seen.add(lw)
            keywords.append(w)
    return keywords


# =====================================================
# --------- Text Chunking with LangChain -------------
# =====================================================

def chunk_text(text: str,
               chunk_size: int = 1000,
               chunk_overlap: int = 200) -> List[str]:
    """
    Split text into overlapping chunks using LangChain's splitter.
    Overlap ensures context continuity and no topic is cut mid-way.
    """
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
        separators=["\n\n", "\n", ". ", " ", ""],
        length_function=len,
    )
    chunks = splitter.split_text(text)
    logger.info(f"Text split into {len(chunks)} chunks (size={chunk_size}, overlap={chunk_overlap})")
    return chunks


# =====================================================
# --------- LLM Structured FAQ Generation ------------
# =====================================================

SYSTEM_PROMPT = """You are a keyword extraction and FAQ generation engine for an insurance search system.

YOUR SINGLE MOST IMPORTANT JOB:
Every distinct word, phrase, number, term, clause name, product name, condition, limit,
percentage, duration, or proper noun present in the chunk MUST appear verbatim (exact spelling,
exact casing) in at least one question or question_variation across the entire output.

This is not optional. The output feeds a suggestion engine: if a word exists in the PDF but
not in any question/variation, users who type that word will get zero suggestions.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
STEP 1 — BEFORE WRITING ANY JSON:
Mentally list EVERY meaningful term in the chunk:
  - All product/plan names
  - All numerical values (amounts, percentages, days, years)
  - All clause/condition names
  - All action words specific to this domain
  - All proper nouns, brand names, legal terms
Then ensure each one ends up verbatim in at least one question or variation.
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
STEP 2 — ANSWERABILITY SELF-CHECK (run before writing each FAQ):
For every question and variation you are about to write, ask yourself:
  "Can this question be answered FULLY and SPECIFICALLY using only the
   text in this chunk — without relying on general insurance knowledge,
   assumptions, or information from outside this chunk?"

If the answer is NO → rewrite the question so it scopes to what the chunk
actually says. If the chunk only partially addresses a topic, write the
question narrowly around what IS stated, not what is implied.

A question is chunk-answerable if and only if:
  ✓ The answer's key fact (number, condition, eligibility rule, process
    step, limit, exclusion) is explicitly present in this chunk's text.
  ✗ The answer requires inferring from general domain knowledge.
  ✗ The answer requires information from another section/chunk not shown.
  ✗ The answer is "it depends" with no further detail in the chunk.
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

STRUCTURE:

1. Extract 8-10 distinct topics from this chunk only.

2. Each topic:
   - topic_id: snake_case
   - title: 2–4 words
   - summary: one line
   - faqs: EXACTLY 5 FAQ entries. Not 1, not 3 — always 5. This is enforced by schema.

3. Each FAQ:

   question:
   - Start with What / How / When / Can I / Does / Why
   - MUST contain at least 2–3 exact terms from the document
   - Up to 20 words — completeness beats brevity
   - MUST be answerable solely from this chunk's text

   question_variations (4–5):
   - Each variation MUST introduce at least one NEW keyword not in the main question
   - Use variations to mop up remaining terms not yet covered
   - Together with the main question they must cover ALL terms for that FAQ's topic area
   - Every variation must also be answerable solely from this chunk's text




HARD RULES:
- Copy terms exactly — do not paraphrase "₹5,000 deductible" as "some deductible"
- Do not invent anything not in the chunk
- Numbers, rupee amounts, day counts must appear as written
- Do NOT write questions whose answers depend on information outside this chunk
- Do NOT write questions that can only be answered with "contact us" or "refer to policy"
  unless those exact words appear in the chunk as the stated resolution
- If the chunk does not contain enough detail to write 5 fully answerable FAQs for a topic,
  reduce the scope of each question further (ask about a sub-detail) rather than fabricating

OUTPUT: valid JSON only, no markdown, no explanation.

{
  "topics": [
    {
      "topic_id": "string",
      "title": "string",
      "faqs": [
        {
          "question": "string",
          "question_variations": ["string", "string", "string", "string"],
        }
      ]
    }
  ]
}
"""


def generate_faq_for_chunk(
    chunk: str,
    chunk_index: int,
    total_chunks: int,
    llm: ChatOpenAI,
    extra_prompt: Optional[str] = None
) -> dict:
    """Generate structured FAQ JSON for a single text chunk."""

    logger.info(f"Processing chunk {chunk_index + 1}/{total_chunks} ({len(chunk)} chars)...")

    structured_llm = llm.with_structured_output(StructuredFAQ)

    user_content = f"DOCUMENT CHUNK {chunk_index + 1} of {total_chunks}:\n\n{chunk}"
    if extra_prompt:
        user_content += f"\n\n{extra_prompt.strip()}"

    messages = [
        SystemMessage(content=SYSTEM_PROMPT),
        HumanMessage(content=user_content)
    ]

    response = structured_llm.invoke(messages)
    result = response.model_dump()

    topic_count = len(result.get("topics", []))
    faq_count = sum(len(t.get("faqs", [])) for t in result.get("topics", []))
    logger.info(f"  → Chunk {chunk_index + 1}: {topic_count} topics, {faq_count} FAQs generated")

    return result


# =====================================================
# --------- Merge & Deduplicate FAQs -----------------
# =====================================================

def merge_structured_faqs(faq_list: List[dict]) -> dict:
    """
    Merge multiple StructuredFAQ dicts into one.
    Topics with the same topic_id are merged; their FAQs are combined.
    Duplicate questions (exact match) are removed.
    """
    merged_topics: Dict[str, dict] = {}

    for faq_data in faq_list:
        for topic in faq_data.get("topics", []):
            tid = topic["topic_id"]

            if tid not in merged_topics:
                merged_topics[tid] = {
                    "topic_id": tid,
                    "title": topic["title"],
                    "faqs": []
                }

            existing_questions = {
                f["question"].strip().lower()
                for f in merged_topics[tid]["faqs"]
            }

            for faq in topic.get("faqs", []):
                q = faq["question"].strip().lower()
                if q not in existing_questions:
                    merged_topics[tid]["faqs"].append(faq)
                    existing_questions.add(q)

    merged = {"topics": list(merged_topics.values())}

    total_topics = len(merged["topics"])
    total_faqs = sum(len(t["faqs"]) for t in merged["topics"])
    logger.info(f"Merged result: {total_topics} topics, {total_faqs} total FAQs")

    return merged


# =====================================================
# --------- Post-process: fix related refs -----------
# =====================================================



# =====================================================
# ---------------- Link Helpers -----------------------
# =====================================================

def load_links_from_xlsx(xlsx_path: Path) -> List[Dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise RuntimeError("openpyxl is required to read .xlsx files") from e

    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    start_idx = 0
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    if "link_id" in header or "linkid" in header or "link" in header:
        start_idx = 1

    links = []
    for row in rows[start_idx:]:
        if not row or len(row) < 2:
            continue
        link_id = str(row[0]).strip() if row[0] is not None else ""
        link_url = str(row[1]).strip() if row[1] is not None else ""
        if not link_id or not link_url:
            continue
        links.append({"link_id": link_id, "link_url": link_url})

    return links


def add_links_to_faqs(structured_data: dict, link_map: Dict[str, str]) -> dict:
    for topic in structured_data.get("topics", []):
        for faq in topic.get("faqs", []):
            link_id = faq.get("link_id")
            faq["link_url"] = link_map.get(link_id) if link_id else None
    return structured_data


def enrich_structured_faq(structured_data: dict, product_name: str) -> dict:
    """Add stable faq_ids and normalize question variations."""
    for topic in structured_data.get("topics", []):
        topic_id = topic.get("topic_id", "").strip()
        for idx, faq in enumerate(topic.get("faqs", [])):
            faq["faq_id"] = f"{product_name}_{topic_id}_{idx}"

            seen = set()
            normalized = []
            for v in faq.get("question_variations", []):
                v = v.strip()
                if v and v not in seen:
                    normalized.append(v)
                    seen.add(v)
            faq["question_variations"] = normalized

    return structured_data


# =====================================================
# ---------------- Vector Store Embed -----------------
# =====================================================

def embed_and_store(structured_data: dict,
                    product_name: str,
                    question_vector_store: VectorStore,
                    rag_vector_store: VectorStore,
                    raw_chunks: List[str] = None):

    logger.info("Preparing documents for embedding...")

    question_documents = []
    rag_documents = []

    for topic in structured_data["topics"]:
        topic_id = topic["topic_id"]

        for idx, faq in enumerate(topic["faqs"]):
            faq_id = faq.get("faq_id", f"{product_name}_{topic_id}_{idx}")
            question_variations = faq.get("question_variations", [])
            link_id = faq.get("link_id")

            # Embed every question + every variation separately for suggestion engine
            all_questions = [faq["question"]] + question_variations
            for q_idx, question_text in enumerate(all_questions):
                question_documents.append({
                    "id": f"{faq_id}_q{q_idx}",
                    "text": question_text.strip(),
                    "metadata": {
                        "product": product_name,
                        "topic_id": topic_id,
                        "topic_name": topic.get("title", ""),
                        "faq_id": faq_id,
                        "question": question_text.strip(),
                        "question_type": "original" if q_idx == 0 else "variation",
                        "question_index": q_idx
                    }
                })



            if raw_chunks:
                rag_documents = [
                    {
                        "id": f"{product_name}_chunk_{i}",
                        "text": chunk.strip(),
                        "metadata": {
                            "product": product_name,
                            "chunk_index": i,
                        }
                    }
                    for i, chunk in enumerate(raw_chunks)
                ]
    
    question_vector_store.add_structured_documents(question_documents)
    rag_vector_store.add_structured_documents(rag_documents)

    total_q = len(question_documents)
    total_r = len(rag_documents)
    logger.info(f"Embedded {total_q} question docs and {total_r} RAG docs into vector stores.")


# =====================================================
# ---------------------- MAIN -------------------------
# =====================================================

def run_ingestion(
    pdf_path: Path,
    product_name: str,
    reset_vector_store: bool = False,
    chunk_size: int = 1000,
    chunk_overlap: int = 200,
    resume_from_checkpoint: bool = False,
) -> None:
    """
    Chunked structured ingestion runner.

    resume_from_checkpoint=True:
      Skips PDF extraction + all LLM calls. Loads the previously saved
      logs/{product_name}_structured_faq.json and runs only the embedding
      step. Use this when embedding failed after JSON was already generated.
    """

    logger.add("logs/structured_ingestion.log", rotation="1 day")
    logger.info("Starting chunked structured ingestion...")
    logger.info(f"PDF: {pdf_path} | Product: {product_name} | resume={resume_from_checkpoint}")

    checkpoint_path = Path("logs") / f"{product_name}_structured_faq.json"

    # ── Vector stores ──────────────────────────────────────────────────────────
    question_vector_store = VectorStore(
        collection_name=settings.chroma_collection_name_questions,
        persist_directory=settings.chroma_persist_directory,
        embedding_model=settings.embedding_model
    )
    rag_vector_store = VectorStore(
        collection_name=settings.chroma_collection_name_rag,
        persist_directory=settings.chroma_persist_directory,
        embedding_model=settings.embedding_model
    )

    if reset_vector_store:
        logger.warning("Resetting vector stores...")
        question_vector_store.reset()
        rag_vector_store.reset()

    try:
        # ══════════════════════════════════════════════════════════════════════
        # RESUME PATH — skip LLM entirely, load existing JSON checkpoint
        # ══════════════════════════════════════════════════════════════════════
        if resume_from_checkpoint:
            if not checkpoint_path.exists():
                raise FileNotFoundError(
                    f"No checkpoint found at {checkpoint_path}. "
                    "Run without resume_from_checkpoint=True first."
                )
            logger.info(f"Resuming from checkpoint: {checkpoint_path}")
            with open(checkpoint_path, "r", encoding="utf-8") as f:
                merged_faq = json.load(f)
            topic_count = len(merged_faq.get("topics", []))
            faq_count   = sum(len(t["faqs"]) for t in merged_faq["topics"])
            logger.info(f"Loaded checkpoint: {topic_count} topics, {faq_count} FAQs.")

        # ══════════════════════════════════════════════════════════════════════
        # FULL PATH — extract → chunk → LLM → merge → save checkpoint
        # ══════════════════════════════════════════════════════════════════════
        else:
            if not pdf_path.exists():
                raise FileNotFoundError(f"PDF file not found: {pdf_path}")

            # ── Load optional link metadata ────────────────────────────────
            xlsx_candidates = list(Path(settings.data_dir).glob("*.xlsx"))
            link_items: List[Dict[str, str]] = []
            if xlsx_candidates:
                link_items = load_links_from_xlsx(xlsx_candidates[0])
                logger.info(f"Loaded {len(link_items)} link entries from XLSX.")

            link_map = {
                li["link_id"]: li["link_url"]
                for li in link_items
                if li.get("link_id") and li.get("link_url")
            }

            extra_prompt = None
            if link_items:
                link_ids = ", ".join([li["link_id"] for li in link_items])
                extra_prompt = (
                    "\nLINK METADATA:\n"
                    "You may attach a relevant link_id to an FAQ ONLY if it directly helps the answer.\n"
                    f"Use ONLY these link_ids: {link_ids}\n"
                    "If no link applies, set link_id to null.\n"
                    "Do not invent ids or urls. Only output link_id; link_url must be null.\n"
                )

            # ── Step 1: Extract text ───────────────────────────────────────
            logger.info("Extracting text from PDF...")
            text = extract_full_text(pdf_path)
            logger.info(f"Extracted {len(text)} characters from PDF.")

            # ── Step 2: Chunk ──────────────────────────────────────────────
            chunks = chunk_text(text, chunk_size=chunk_size, chunk_overlap=chunk_overlap)

            # ── Step 3: LLM per chunk ──────────────────────────────────────
            # ── Step 3: LLM per chunk (parallel) ──────────────────────────────
            llm = ChatOpenAI(
                model=settings.openai_model_large,
                temperature=settings.llm_temperature,
                api_key=settings.openai_api_key,
                max_retries=3,
            )

            all_chunk_faqs = [None] * len(chunks)  # preserve order

            with ThreadPoolExecutor(max_workers=2) as executor:
                futures = {
                    executor.submit(
                        generate_faq_for_chunk,
                        chunk=chunk,
                        chunk_index=i,
                        total_chunks=len(chunks),
                        llm=llm,
                        extra_prompt=extra_prompt
                    ): i
                    for i, chunk in enumerate(chunks)
                }
                for future in as_completed(futures):
                    i = futures[future]
                    result = future.result()
                    all_chunk_faqs[i] = result

                    # ── Print chunk output for inspection ─────────────────────
                    print(f"\n{'='*60}")
                    print(f"CHUNK {i + 1}/{len(chunks)}")
                    print(f"{'='*60}")
                    for topic in result.get("topics", []):
                        print(f"\n  📁 [{topic['topic_id']}] {topic['title']}")
                        for faq in topic.get("faqs", []):
                            print(f"     Q: {faq['question']}")
                            for v in faq.get("question_variations", []):
                                print(f"        ~ {v}")
                    print(f"{'='*60}\n")

                    logger.info(f"Chunk {i + 1}/{len(chunks)} completed.",flush=True)

            logger.info(f"All {len(chunks)} chunks processed. Merging...")

            # ── Step 4: Merge + post-process ──────────────────────────────
            merged_faq = merge_structured_faqs(all_chunk_faqs)
            merged_faq = enrich_structured_faq(merged_faq, product_name)
            # merged_faq = add_links_to_faqs(merged_faq, link_map)

            topic_count = len(merged_faq.get("topics", []))
            faq_count   = sum(len(t["faqs"]) for t in merged_faq["topics"])
            logger.info(f"Final merged FAQ: {topic_count} topics, {faq_count} FAQs.")

            if topic_count < 10:
                logger.warning(f"Only {topic_count} topics — consider smaller chunk_size.")

            # ── Step 5: Save checkpoint (always, before embedding) ─────────
            checkpoint_path.parent.mkdir(parents=True, exist_ok=True)
            with open(checkpoint_path, "w", encoding="utf-8") as f:
                json.dump(merged_faq, f, indent=2, ensure_ascii=False)
            logger.info(f"Checkpoint saved → {checkpoint_path}")

        # ══════════════════════════════════════════════════════════════════════
        # EMBED (runs in both full and resume paths)
        # ══════════════════════════════════════════════════════════════════════
        embed_and_store(
            merged_faq,
            product_name,
            question_vector_store,
            rag_vector_store,
            raw_chunks=chunks
        )

        logger.info("Ingestion completed successfully.")
        print(f"\n✅ Done! {topic_count} topics | {faq_count} FAQs embedded.")

    except Exception as e:
        logger.error(f"Ingestion failed: {e}")
        print(f"\n❌ Error: {e}")
        raise


# if __name__ == "__main__":
#     run_ingestion(
#         pdf_path=Path(settings.pdf_dir),
#         product_name=settings.default_product_name,
#         reset_vector_store=settings.reset_vector_store,
#         chunk_size=2000,
#         chunk_overlap=300,
#         resume_from_checkpoint=True,   # ← set True to skip LLM and re-embed only
#     )


